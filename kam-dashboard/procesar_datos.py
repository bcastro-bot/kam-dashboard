"""
procesar_datos.py — KAM Dashboard Multi-KAM
Genera docs/data.json con datos de TODOS los KAMs.

Lógica de clientes:
- Fuente de verdad: hoja 'Clientes' (columna VENDEDOR ACTUAL + PANEL + RUT)
- Cruce: RUT limpio (solo dígitos) entre hoja Clientes y hoja Facturas
- Clientes activos del KAM: compraron desde Ene 2025 a la fecha
- Clientes sin compra desde Ene 2025: se asignan a KAM 'PERDIDOS'
- Forecast = mismo mes año anterior × FORECAST_GROWTH
"""

import sys
import json
import re as _re
from pathlib import Path

try:
    import pandas as pd
except ImportError:
    print("ERROR: pip install pandas openpyxl"); sys.exit(1)

# ── Configuración ──────────────────────────────────────────────────────────────
SHEET_NAMES     = ["CUADRATURAS", "OTROS_CUADRATURADOCUMENTOSINGRE"]
REF_DATE        = pd.Timestamp("2026-08-01")
CORTE_ACTIVO    = "2025-01"   # Desde este mes para considerar cliente activo del KAM
OUTPUT_FILE     = Path("docs/data.json")
FORECAST_GROWTH = 1.35
VALID_KAMS      = ["BC", "BG", "LJ", "CF", "AA", "DA", "SC", "EC", "SG", "BACK", "PERDIDOS"]

ALL_POSSIBLE_MONTHS = [
    "2024-01","2024-02","2024-03","2024-04","2024-05","2024-06",
    "2024-07","2024-08","2024-09","2024-10","2024-11","2024-12",
    "2025-01","2025-02","2025-03","2025-04","2025-05","2025-06",
    "2025-07","2025-08","2025-09","2025-10","2025-11","2025-12",
    "2026-01","2026-02","2026-03","2026-04","2026-05","2026-06",
    "2026-07","2026-08","2026-09","2026-10","2026-11","2026-12",
]

MONTH_LABELS = {
    "2024-01":"Ene 24","2024-02":"Feb 24","2024-03":"Mar 24","2024-04":"Abr 24",
    "2024-05":"May 24","2024-06":"Jun 24","2024-07":"Jul 24","2024-08":"Ago 24",
    "2024-09":"Sep 24","2024-10":"Oct 24","2024-11":"Nov 24","2024-12":"Dic 24",
    "2025-01":"Ene 25","2025-02":"Feb 25","2025-03":"Mar 25","2025-04":"Abr 25",
    "2025-05":"May 25","2025-06":"Jun 25","2025-07":"Jul 25","2025-08":"Ago 25",
    "2025-09":"Sep 25","2025-10":"Oct 25","2025-11":"Nov 25","2025-12":"Dic 25",
    "2026-01":"Ene 26","2026-02":"Feb 26","2026-03":"Mar 26","2026-04":"Abr 26",
    "2026-05":"May 26","2026-06":"Jun 26","2026-07":"Jul 26",
    "2026-08":"Ago 26*","2026-09":"Sep 26*","2026-10":"Oct 26*",
    "2026-11":"Nov 26*","2026-12":"Dic 26*",
}

FORECAST_BASE = {
    "2026-08":"2025-08","2026-09":"2025-09","2026-10":"2025-10",
    "2026-11":"2025-11","2026-12":"2025-12",
}

# ── Helpers ────────────────────────────────────────────────────────────────────
def clean_rut(r):
    """Extrae solo los dígitos del RUT."""
    return _re.sub(r'[^0-9]', '', str(r))

def categorize_product(producto):
    if pd.isna(producto): return "Sin clasificar"
    p = str(producto).strip().lower()
    if _re.match(r"^gif?t?t?\s*card\s*-?\s*po\.", p) or "gift card -" in p or "gift card po" in p or "gitt card" in p: return "Gift Card"
    if "gift card" in p or "giftcard" in p or "gc rewards" in p or "gc market" in p: return "Gift Card"
    if "puntos" in p: return "Puntos"
    if "software" in p: return "Software"
    if "agencia" in p: return "Agencia"
    if "comisión" in p or "comision" in p: return "Comisión"
    if "efectivo" in p: return "Efectivo"
    if "fee" in p or "saas" in p: return "Fee SaaS"
    if "plataforma" in p: return "Plataforma"
    if "rebaja" in p: return "Ajuste"
    return "Otros"

def fmt_short(m):
    if not m: return ""
    names = ["","Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    return names[int(m.split("-")[1])]

def pct(a, b):
    if not b or b == 0: return None
    return round((a - b) / b * 100, 1)

# ── Carga ──────────────────────────────────────────────────────────────────────
def load_data(paths):
    """
    Carga facturas y hoja Clientes.
    Cruce por RUT (solo dígitos).
    Retorna df con columnas: Kam, Cliente (nombre PANEL), Rut_clean, Precio, Mes, Categoria
    """
    # 1. Leer hoja Clientes (fuente de verdad)
    clientes_df = None
    for p in paths:
        p = Path(p)
        if not p.exists(): continue
        import openpyxl
        wb = openpyxl.load_workbook(p, read_only=True)
        if "Clientes" in wb.sheetnames:
            clientes_df = pd.read_excel(p, sheet_name="Clientes")
            clientes_df.columns = [c.strip() for c in clientes_df.columns]
            clientes_df['RUT_clean'] = clientes_df['RUT'].apply(clean_rut)
            clientes_df['PANEL'] = clientes_df['PANEL'].astype(str).str.strip()
            clientes_df['VENDEDOR ACTUAL'] = clientes_df['VENDEDOR ACTUAL'].astype(str).str.strip()
            print(f"  ✓ Hoja Clientes: {len(clientes_df)} registros")
            print(f"    KAMs disponibles: {clientes_df['VENDEDOR ACTUAL'].value_counts().head(10).to_dict()}")
            break

    if clientes_df is None:
        print("  ⚠ No se encontró hoja Clientes — usando KAM de facturas")

    # Mapa: RUT_clean → (PANEL, VENDEDOR_ACTUAL)
    rut_to_panel = {}
    rut_to_kam   = {}
    if clientes_df is not None:
        for _, row in clientes_df.iterrows():
            rut = row['RUT_clean']
            if rut and rut != 'nan':
                rut_to_panel[rut] = row['PANEL']
                rut_to_kam[rut]   = row['VENDEDOR ACTUAL']

    # 2. Leer facturas
    frames = []
    for p in paths:
        p = Path(p)
        if not p.exists():
            print(f"  ✗ No encontré: {p}"); continue
        import openpyxl
        wb = openpyxl.load_workbook(p, read_only=True)
        sheet = None
        for s in SHEET_NAMES:
            if s in wb.sheetnames:
                sheet = s; break
        if not sheet:
            print(f"  ⚠ Sin hoja de facturas en {p.name}"); continue

        print(f"  ✓ Cargando facturas: {p.name} (hoja: {sheet})")
        df = pd.read_excel(p, sheet_name=sheet)
        df.columns = [str(c).strip() for c in df.columns]

        # Normalizar precio
        for col in ["Precio","Prrecio","precio","PRECIO"]:
            if col in df.columns:
                df = df.rename(columns={col:"Precio"}); break

        # Normalizar KAM original
        for col in ["Kam","KAM","Codigo Vendedor"]:
            if col in df.columns:
                df = df.rename(columns={col:"Kam_orig"}); break

        df = df[df["Precio"].notna() & (df["Precio"] != 0)].copy()
        df["Fecha Emision"] = pd.to_datetime(df["Fecha Emision"])
        df["Mes"] = df["Fecha Emision"].dt.to_period("M").astype(str)
        df["RUT_clean"] = df["Rut"].apply(clean_rut)

        # Cruce por RUT → obtener nombre PANEL y VENDEDOR ACTUAL
        df["Cliente"] = df["RUT_clean"].map(rut_to_panel).fillna(df.get("Cliente", df.get("RazonSocial","")))
        df["Kam"]     = df["RUT_clean"].map(rut_to_kam).fillna(df.get("Kam_orig",""))

        if "Producto" in df.columns:
            df["Categoria"] = df["Producto"].apply(categorize_product)
        else:
            df["Categoria"] = "Sin clasificar"

        matched = df["RUT_clean"].isin(rut_to_panel).sum()
        print(f"    → {len(df)} registros · {matched} cruzados por RUT ({matched/len(df)*100:.0f}%)")
        frames.append(df[["Kam","Cliente","RUT_clean","Precio","Mes","Categoria"]])

    if not frames:
        print("ERROR: No se cargaron facturas."); sys.exit(1)

    combined = pd.concat(frames, ignore_index=True)

    # 3. Separar clientes activos vs perdidos
    # Activo del KAM = compró desde CORTE_ACTIVO en adelante
    compras_recientes = combined[(combined["Mes"] >= CORTE_ACTIVO) & (combined["Precio"] > 0)]
    ruts_activos = set(compras_recientes["RUT_clean"].unique())

    # Reasignar a PERDIDOS si nunca compraron desde CORTE_ACTIVO
    combined["Kam"] = combined.apply(
        lambda r: r["Kam"] if r["RUT_clean"] in ruts_activos else "PERDIDOS",
        axis=1
    )

    # Filtrar solo KAMs válidos
    combined = combined[combined["Kam"].isin(VALID_KAMS)]

    print(f"\n  Distribución final por KAM:")
    dist = combined[combined["Precio"]>0].groupby("Kam")["RUT_clean"].nunique().sort_values(ascending=False)
    for kam, n in dist.items():
        rev = combined[(combined["Kam"]==kam) & (combined["Precio"]>0)]["Precio"].sum()
        print(f"    {kam}: {n} clientes · ${rev:,.0f}")

    return combined

# ── Dataset por KAM ────────────────────────────────────────────────────────────
def build_kam_data(df_kam):
    # Agrupar por nombre de cliente (PANEL)
    monthly = df_kam.groupby(["Cliente","Mes"])["Precio"].sum().reset_index()
    real_months = sorted(monthly["Mes"].unique().tolist())
    hist_months = [m for m in ALL_POSSIBLE_MONTHS if m in real_months]
    forecast_months = sorted([m for m in FORECAST_BASE if m not in real_months])
    all_months = hist_months + forecast_months

    years = sorted(set(m[:4] for m in hist_months))
    months_by_year = {y:[m for m in hist_months if m.startswith(y)] for y in years}
    cur_y = max(years)
    prev_y = str(int(cur_y)-1)
    yoy_pairs = [(prev_y+"-"+m[5:], m) for m in months_by_year.get(cur_y,[])
                 if prev_y+"-"+m[5:] in hist_months]

    clients = {}
    for client, grp in monthly.groupby("Cliente"):
        vd = dict(zip(grp["Mes"], grp["Precio"]))
        hist_vals   = [int(vd.get(m,0)) for m in hist_months]
        fc_vals     = [int(vd.get(FORECAST_BASE[m],0)*FORECAST_GROWTH) for m in forecast_months]
        total       = sum(hist_vals)
        active      = [hist_months[i] for i,v in enumerate(hist_vals) if v>0]
        freq        = len(active)
        last_m      = active[-1] if active else None
        first_m     = active[0]  if active else None
        days        = int((REF_DATE - pd.Timestamp(last_m+"-28")).days) if last_m else 999
        rev_by_year = {y: sum(vd.get(m,0) for m in months_by_year.get(y,[])) for y in years}

        months_cur = months_by_year.get(cur_y,[])
        ytd_by_year = {}
        for y in years:
            ytd_by_year[y] = sum(vd.get(y+"-"+m[5:],0) for m in months_cur if y+"-"+m[5:] in hist_months)
        ytd_prev = ytd_by_year.get(prev_y,0)
        ytd_cur  = ytd_by_year.get(cur_y,0)

        avg_cur = int(sum(vd.get(m,0) for m in months_cur)/len(months_cur)) if months_cur else 0

        last3 = hist_months[-3:]
        has_recent = any(vd.get(m,0)>0 for m in last3)
        has_prev   = rev_by_year.get(prev_y,0)>0
        mid        = hist_months[max(0,len(hist_months)-7):-3]
        has_mid    = any(vd.get(m,0)>0 for m in mid)

        if has_recent:
            status = "nuevo" if not has_prev else "activo"
        elif rev_by_year.get(cur_y,0)>0 or has_mid:
            status = "en riesgo"
        else:
            status = "churn"

        clients[client] = {
            "total": total, "freq": freq, "ultima": last_m, "dias": days, "status": status,
            "rev_by_year": {y:int(v) for y,v in rev_by_year.items()},
            "rev_2025": int(rev_by_year.get("2025",0)),
            "rev_2026": int(rev_by_year.get("2026",0)),
            "ytd24": int(ytd_by_year.get("2024",0)),
            "ytd25": int(ytd_by_year.get("2025",0)),
            "ytd26": int(ytd_by_year.get("2026",0)),
            "ytd_pct": pct(ytd_cur, ytd_prev),
            "avg_2026": avg_cur,
            "vals": [round(v/1000) for v in hist_vals] + [round(v/1000) for v in fc_vals],
            "hist_count": len(hist_months),
            "forecast_vals": [round(v/1000) for v in fc_vals],
            "forecast_total": round(sum(fc_vals)/1000),
        }

    # Cartera
    macro = {m: int(monthly[monthly["Mes"]==m]["Precio"].sum()) for m in hist_months}
    for m in forecast_months:
        macro[m] = int(monthly[monthly["Mes"]==FORECAST_BASE[m]]["Precio"].sum()*FORECAST_GROWTH)

    mh = [macro.get(m,0) for m in hist_months]
    mf = [macro.get(m,0) for m in forecast_months]
    ytd_c = {y: sum(macro.get(y+"-"+m[5:],0) for m in months_by_year.get(cur_y,[]) if y+"-"+m[5:] in hist_months) for y in years}

    cartera = {
        "total": sum(mh),
        "vals":  [round(v/1000) for v in mh] + [round(v/1000) for v in mf],
        "hist_count": len(hist_months),
        "forecast_vals":  [round(v/1000) for v in mf],
        "forecast_total": round(sum(mf)/1000),
        "rev_by_year": {y: int(sum(macro.get(m,0) for m in months_by_year.get(y,[]))) for y in years},
        "mensual": [{"lbl":fmt_short(p[0]),"v25":macro.get(p[0],0),"v26":macro.get(p[1],0),"pct":pct(macro.get(p[1],0),macro.get(p[0],0))} for p in yoy_pairs],
        "trimestral": [{"lbl":"Q1","v25":int(sum(macro.get(m,0) for m in months_by_year.get(prev_y,[]) if m.endswith(("-01","-02","-03")))),"v26":int(sum(macro.get(m,0) for m in months_by_year.get(cur_y,[]) if m.endswith(("-01","-02","-03")))),"pct":None}],
        "ytd25": int(ytd_c.get(prev_y,0)), "ytd26": int(ytd_c.get(cur_y,0)),
        "ytd_pct": pct(ytd_c.get(cur_y,0), ytd_c.get(prev_y,0)),
        "years": years,
    }

    # Tickets por mes
    tix = df_kam[df_kam["Precio"]>0].groupby("Mes").agg(tickets=("Precio","count"),revenue=("Precio","sum")).reset_index()
    tickets_por_mes = {}
    for m in hist_months:
        row = tix[tix["Mes"]==m]
        if len(row):
            t = int(row["tickets"].values[0]); r = int(row["revenue"].values[0])
            tickets_por_mes[m] = {"tickets":t,"revenue":r,"ticket_prom":int(r/t) if t else 0,"label":MONTH_LABELS.get(m,m)}
        else:
            tickets_por_mes[m] = {"tickets":0,"revenue":0,"ticket_prom":0,"label":MONTH_LABELS.get(m,m)}

    # Clientes activos por mes
    mn = df_kam.groupby(["Cliente","Mes"])["Precio"].sum().reset_index()
    mn = mn[mn["Precio"]>0]
    clientes_por_mes = {m: {"count":int(mn[mn["Mes"]==m]["Cliente"].nunique()),"revenue":int(mn[mn["Mes"]==m]["Precio"].sum()),"label":MONTH_LABELS.get(m,m)} for m in hist_months}

    return {
        "months": all_months,
        "month_labels": [MONTH_LABELS.get(m,m) for m in all_months],
        "hist_count": len(hist_months),
        "forecast_months": forecast_months,
        "years": years,
        "cartera": cartera,
        "clients": clients,
        "tickets_por_mes": tickets_por_mes,
        "clientes_por_mes": clientes_por_mes,
    }

# ── Main ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    files = sys.argv[1:] if len(sys.argv)>1 else list(Path(".").glob("*.xlsx")) + list(Path("data").glob("*.xlsx"))
    if not files:
        print("USO: python procesar_datos.py archivo.xlsx"); sys.exit(1)

    print(f"\n{'='*54}")
    print(f"  KAM Dashboard — Multi-KAM (2024–2026)")
    print(f"  Corte activos: desde {CORTE_ACTIVO}")
    print(f"{'='*54}\n")

    df = load_data(files)
    kams_presentes = [k for k in VALID_KAMS if k in df["Kam"].values]

    output = {
        "generated_at": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"),
        "kams": kams_presentes,
        "corte_activo": CORTE_ACTIVO,
    }

    print(f"\n  Procesando KAMs: {kams_presentes}\n")
    for kam in kams_presentes:
        df_kam = df[df["Kam"]==kam].copy()
        print(f"  {kam}: {df_kam['Cliente'].nunique()} clientes")
        output[kam] = build_kam_data(df_kam)

    # Productos por categoría
    all_months_sorted = sorted(df["Mes"].unique().tolist())
    productos_data = {}
    for kam in kams_presentes:
        df_kam = df[df["Kam"]==kam]
        if not len(df_kam): continue
        cat_m = df_kam.groupby(["Categoria","Mes"])["Precio"].sum().reset_index()
        cat_c = df_kam[df_kam["Precio"]>0].groupby("Categoria")["Precio"].agg(["sum","count"]).reset_index()
        cat_data = {}
        for cat in sorted(df_kam["Categoria"].unique()):
            vals = [int(cat_m[(cat_m["Categoria"]==cat)&(cat_m["Mes"]==m)]["Precio"].values[0]) if len(cat_m[(cat_m["Categoria"]==cat)&(cat_m["Mes"]==m)]) else 0 for m in all_months_sorted]
            total = sum(vals)
            if total > 0:
                cr = cat_c[cat_c["Categoria"]==cat]
                n  = int(cr["count"].values[0]) if len(cr) else 0
                rp = int(cr["sum"].values[0]) if len(cr) else total
                cat_data[cat] = {"vals":vals,"total":total,"tickets":n,"ticket_prom":int(rp/n) if n else 0}
        productos_data[kam] = cat_data

    output["productos_data"]   = productos_data
    output["productos_months"] = all_months_sorted

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE,"w",encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, separators=(",",":"))

    kb = OUTPUT_FILE.stat().st_size/1024
    print(f"\n  ✓ {OUTPUT_FILE} · {kb:.0f} KB")
    print(f"  ✓ Años: {output[kams_presentes[0]]['years']}")
    print(f"  ✓ Corte activos: desde {CORTE_ACTIVO}\n")
    for kam in kams_presentes:
        t = output[kam]["cartera"]["total"]
        n = len(output[kam]["clients"])
        print(f"    {kam}: {n} clientes · ${t:,.0f}")
    print()
